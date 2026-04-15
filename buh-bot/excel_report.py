"""
Генератор Excel-отчётов.
Создаёт .xlsx файл с двумя листами:
  1. Доп. работы по клиентам за месяц
  2. KPI бухгалтеров за месяц
"""
import os
from datetime import datetime
from typing import List

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

import database as db

# ─── Цвета ────────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill('solid', fgColor='1F4E79')  # тёмно-синий
SECTION_FILL = PatternFill('solid', fgColor='D6E4F0')  # светло-голубой
TOTAL_FILL   = PatternFill('solid', fgColor='E2EFDA')  # светло-зелёный
RED_FILL     = PatternFill('solid', fgColor='FFD7D7')  # красный
WHITE        = PatternFill('solid', fgColor='FFFFFF')

HEADER_FONT  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
BOLD_FONT    = Font(name='Calibri', bold=True, size=10)
NORMAL_FONT  = Font(name='Calibri', size=10)

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)


def _set_header(ws, row: int, cols: list):
    for col_idx, title in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _style_cell(cell, font=None, fill=None, align=None, fmt=None):
    cell.font = font or NORMAL_FONT
    if fill:
        cell.fill = fill
    cell.alignment = align or LEFT
    cell.border = THIN_BORDER
    if fmt:
        cell.number_format = fmt


def fmt_date(iso_date: str) -> str:
    try:
        return datetime.strptime(iso_date[:10], '%Y-%m-%d').strftime('%d.%m.%Y')
    except Exception:
        return iso_date or '—'


# ─── Лист 1: Доп. работы ──────────────────────────────────────────────────────

def _build_works_sheet(wb, year: int, month: int):
    ws = wb.active
    ws.title = 'Доп. работы'

    month_name = datetime(year, month, 1).strftime('%B %Y')

    # Заголовок отчёта
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f'ОТЧЁТ ПО ДОПОЛНИТЕЛЬНЫМ РАБОТАМ — {month_name.upper()}'
    title_cell.font = Font(name='Calibri', bold=True, size=13, color='1F4E79')
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 25

    ws.row_dimensions[2].height = 5  # отступ

    # Заголовки колонок
    headers = ['Клиент', 'Бухгалтер', 'Дата', 'Тип работы', 'Описание', 'Часы', 'Сумма, ₽', 'Примечание']
    _set_header(ws, 3, headers)
    ws.row_dimensions[3].height = 20

    works = db.get_additional_works_for_month(year, month)

    if not works:
        ws.cell(row=4, column=1, value='Доп. работ за период не зафиксировано')
        return ws

    # Группируем по компании
    by_company: dict = {}
    for w in works:
        co = w['company_name']
        if co not in by_company:
            by_company[co] = []
        by_company[co].append(w)

    row = 4
    grand_hours = 0.0
    grand_amount = 0.0

    for co_name, ws_list in by_company.items():
        # Строка-заголовок компании
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws.cell(row=row, column=1, value=f'  🏢  {co_name}')
        cell.font = BOLD_FONT
        cell.fill = SECTION_FILL
        cell.alignment = LEFT
        cell.border = THIN_BORDER
        ws.row_dimensions[row].height = 18
        row += 1

        co_hours = 0.0
        co_amount = 0.0

        for w in ws_list:
            data = [
                co_name,
                w['accountant_name'] or '—',
                fmt_date(w['work_date']),
                w['work_type'],
                w['description'],
                w['hours'] or 0,
                w['amount'] or 0,
                '',
            ]
            for col_idx, val in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                _style_cell(cell, fill=WHITE,
                            align=CENTER if col_idx in (3, 6, 7) else LEFT,
                            fmt='#,##0.00' if col_idx == 7 else None)
            co_hours += w['hours'] or 0
            co_amount += w['amount'] or 0
            ws.row_dimensions[row].height = 15
            row += 1

        # Итого по компании
        for col_idx in range(1, 9):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = TOTAL_FILL
            cell.border = THIN_BORDER
        ws.cell(row=row, column=1, value='ИТОГО по компании').font = BOLD_FONT
        ws.cell(row=row, column=1).fill = TOTAL_FILL
        ws.cell(row=row, column=1).alignment = LEFT
        ws.cell(row=row, column=6, value=co_hours).font = BOLD_FONT
        ws.cell(row=row, column=6).fill = TOTAL_FILL
        ws.cell(row=row, column=6).alignment = CENTER
        ws.cell(row=row, column=7, value=co_amount).font = BOLD_FONT
        ws.cell(row=row, column=7).fill = TOTAL_FILL
        ws.cell(row=row, column=7).alignment = CENTER
        ws.cell(row=row, column=7).number_format = '#,##0.00'
        ws.row_dimensions[row].height = 16
        row += 1

        grand_hours += co_hours
        grand_amount += co_amount

    # Итого общий
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1, value='ИТОГО ЗА МЕСЯЦ').font = Font(name='Calibri', bold=True, size=11, color='1F4E79')
    ws.cell(row=row, column=1).alignment = LEFT
    ws.cell(row=row, column=6, value=grand_hours).font = Font(name='Calibri', bold=True, size=11)
    ws.cell(row=row, column=6).alignment = CENTER
    ws.cell(row=row, column=7, value=grand_amount).font = Font(name='Calibri', bold=True, size=11)
    ws.cell(row=row, column=7).alignment = CENTER
    ws.cell(row=row, column=7).number_format = '#,##0.00'
    for col_idx in range(1, 9):
        ws.cell(row=row, column=col_idx).fill = PatternFill('solid', fgColor='BDD7EE')
        ws.cell(row=row, column=col_idx).border = THIN_BORDER

    # Ширина колонок
    col_widths = [25, 18, 12, 22, 40, 8, 14, 15]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return ws


# ─── Лист 2: KPI бухгалтеров ──────────────────────────────────────────────────

def _build_kpi_sheet(wb, year: int, month: int):
    ws = wb.create_sheet(title='KPI бухгалтеров')
    month_name = datetime(year, month, 1).strftime('%B %Y')

    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = f'KPI БУХГАЛТЕРОВ — {month_name.upper()}'
    title_cell.font = Font(name='Calibri', bold=True, size=13, color='1F4E79')
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 25

    headers = [
        'Бухгалтер', 'Компаний', 'Дедлайнов', 'Выполнено',
        'Просрочено', '% выполнения', 'Ошибок', 'Коэф. ошибок',
        'Доп.часы', 'Доп.сумма, ₽'
    ]
    _set_header(ws, 3, headers)
    ws.row_dimensions[3].height = 22

    stats = db.get_accountant_stats_full(year, month)

    row = 4
    for s in stats:
        total = s['total_deadlines'] or 0
        done = s['done_deadlines'] or 0
        overdue = s['overdue_deadlines'] or 0
        errors = s['error_count'] or 0
        pct = round(done / total * 100, 1) if total > 0 else 0
        err_coef = round(errors / total * 100, 1) if total > 0 else 0

        data = [
            s['name'],
            s['company_count'] or 0,
            total,
            done,
            overdue,
            pct,
            errors,
            err_coef,
            s['extra_hours'] or 0,
            s['extra_amount'] or 0,
        ]
        for col_idx, val in enumerate(data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = CENTER if col_idx > 1 else LEFT
            cell.font = NORMAL_FONT
            if col_idx == 5 and overdue > 0:
                cell.fill = RED_FILL
                cell.font = Font(name='Calibri', bold=True, size=10, color='C00000')
            if col_idx == 7 and errors > 0:
                cell.fill = RED_FILL
                cell.font = Font(name='Calibri', bold=True, size=10, color='C00000')
            if col_idx in (6, 8):
                cell.number_format = '0.0"%"'
            if col_idx == 10:
                cell.number_format = '#,##0.00'
        ws.row_dimensions[row].height = 16
        row += 1

    col_widths = [20, 10, 11, 11, 11, 13, 9, 13, 10, 15]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return ws


# ─── Публичная функция ────────────────────────────────────────────────────────

def generate_monthly_report(year: int = None, month: int = None) -> str:
    """
    Генерирует Excel-отчёт, сохраняет в data_storage/ и возвращает путь к файлу.
    """
    now = datetime.now()
    if year is None:
        year = now.year
    if month is None:
        month = now.month

    wb = openpyxl.Workbook()
    _build_works_sheet(wb, year, month)
    _build_kpi_sheet(wb, year, month)

    os.makedirs('data_storage', exist_ok=True)
    filename = f'data_storage/report_{year}_{month:02d}.xlsx'
    wb.save(filename)
    return filename
