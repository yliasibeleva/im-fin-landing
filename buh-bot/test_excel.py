import sys, io
sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, 'D:\\РС\\Визитка\\buh-bot')
import database as db
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

today = date.today()
companies_raw = db.get_all_companies()
kpi_raw       = db.get_accountant_stats_full(today.year, today.month)

wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = 'Компании'

hdr_font  = Font(bold=True, color='FFFFFF', size=10)
hdr_fill  = PatternFill('solid', fgColor='1A3A5C')
hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
fill_a = PatternFill('solid', fgColor='EFF4FB')
fill_b = PatternFill('solid', fgColor='DDEAF8')

headers = ['ID', 'Название', 'СНО', 'Тип', 'Бухгалтер', 'Сотрудники', 'Воинский учёт']
for ci, h in enumerate(headers, 1):
    c = ws1.cell(row=1, column=ci, value=h)
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align

col_widths = [6, 36, 10, 8, 24, 12, 14]
for ci, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(ci)].width = w

companies_list = list(companies_raw)
for ri, c in enumerate(companies_list, 2):
    cd = dict(c)
    row_fill = fill_a if ri % 2 == 0 else fill_b
    vals = [cd['id'], cd['name'], cd['tax_system'], cd['org_type'],
            cd.get('accountant_name') or '',
            'Да' if cd.get('has_employees') else 'Нет',
            'Да' if cd.get('has_military') else 'Нет']
    for ci2, val in enumerate(vals, 1):
        cell = ws1.cell(row=ri, column=ci2, value=val)
        cell.fill = row_fill
        cell.alignment = Alignment(vertical='center')

ws1.auto_filter.ref = f'A1:G{len(companies_list)+1}'

ws2 = wb.create_sheet('KPI бухгалтеров')
kpi_headers = ['Бухгалтер', 'Компаний', 'Дедлайнов', 'Выполнено', 'Просрочено', '% исп.', 'Ошибок', 'Доп.часы', 'Доп.сумма']
for ci, h in enumerate(kpi_headers, 1):
    c = ws2.cell(row=1, column=ci, value=h)
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align

for ri, s in enumerate(kpi_raw, 2):
    sd = dict(s)
    total = sd.get('total_deadlines') or 0
    done  = sd.get('done_deadlines') or 0
    pct   = round(done / total * 100) if total > 0 else 0
    row_fill = fill_a if ri % 2 == 0 else fill_b
    vals = [sd['name'], sd.get('company_count') or 0, total, done,
            sd.get('overdue_deadlines') or 0, f'{pct}%',
            sd.get('error_count') or 0, sd.get('extra_hours') or 0, sd.get('extra_amount') or 0]
    for ci2, val in enumerate(vals, 1):
        cell = ws2.cell(row=ri, column=ci2, value=val)
        cell.fill = row_fill
        cell.alignment = Alignment(vertical='center')

buf = io.BytesIO()
wb.save(buf)
buf.seek(0)
print(f'Excel OK: {len(buf.read())} байт')
