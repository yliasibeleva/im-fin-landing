"""
Веб-дашборд для руководителя.
Запуск: python web_app.py
Открыть: http://localhost:8000
Логин: admin / пароль из .env (WEB_PASSWORD)
"""
import asyncio
import io
import os
import secrets
from datetime import date
try:
    import httpx as _httpx
    _HTTPX_OK = True
except ImportError:
    _HTTPX_OK = False

import uvicorn
from dotenv import load_dotenv
from fastapi import Depends, FastAPI, Form, HTTPException, Request, status
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from fastapi.templating import Jinja2Templates
from typing import Annotated, Optional

import database as db

load_dotenv()

db.init_db()  # применяем миграции при старте

WEB_PASSWORD = os.getenv('WEB_PASSWORD', 'admin')
WEB_PORT     = int(os.getenv('WEB_PORT', '8000'))

TG_TOKEN = os.getenv('TELEGRAM_BOT_TOKEN', '')
TG_CHAT  = os.getenv('TELEGRAM_CHAT_ID', '')


async def tg(text: str) -> None:
    """Отправляет уведомление в Telegram. Молчит если токен не настроен."""
    if not TG_TOKEN or not TG_CHAT or not _HTTPX_OK:
        return
    try:
        async with _httpx.AsyncClient(timeout=5) as client:
            await client.post(
                f'https://api.telegram.org/bot{TG_TOKEN}/sendMessage',
                json={'chat_id': TG_CHAT, 'text': text, 'parse_mode': 'HTML'},
            )
    except Exception:
        pass

app = FastAPI(title='Империя Финанс — Дашборд')
templates = Jinja2Templates(directory='templates')
security = HTTPBasic()

# Фильтр для форматирования рублей
templates.env.filters['rub'] = lambda v: f"{int(v or 0):,}".replace(',', '\u00a0')


# ─── Авторизация ──────────────────────────────────────────────────────────────

def require_auth(credentials: Annotated[HTTPBasicCredentials, Depends(security)]):
    ok = secrets.compare_digest(
        credentials.password.encode('utf-8'),
        WEB_PASSWORD.encode('utf-8')
    )
    if not ok:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            headers={'WWW-Authenticate': 'Basic'},
        )


# ─── Утилиты ──────────────────────────────────────────────────────────────────

def fmt_date(iso: str) -> str:
    if not iso:
        return '—'
    try:
        return date.fromisoformat(iso[:10]).strftime('%d.%m.%Y')
    except ValueError:
        return iso


def days_left(iso: str) -> int:
    try:
        return (date.fromisoformat(iso[:10]) - date.today()).days
    except ValueError:
        return 0


# ─── Маршруты ─────────────────────────────────────────────────────────────────

@app.get('/', response_class=HTMLResponse)
async def dashboard(
    request: Request, _=Depends(require_auth),
    q: Optional[str] = None,
    acc: Optional[str] = None,
    op: Optional[str] = None,
    tax: Optional[str] = None,
    emp: Optional[str] = None,
    mil: Optional[str] = None,
):
    today = date.today()

    companies_raw   = await asyncio.to_thread(db.get_all_companies)
    accountants_raw = await asyncio.to_thread(db.get_all_accountants)

    companies = [dict(c) for c in companies_raw]

    # Применяем фильтры таблицы компаний
    if q:
        ql = q.lower()
        companies = [c for c in companies if
            ql in c['name'].lower() or
            ql in (c.get('accountant_name') or '').lower() or
            ql in (c.get('operator_name') or '').lower() or
            ql in (c.get('payroll_accountant_name') or '').lower() or
            ql in (c.get('hr_accountant_name') or '').lower()]
    if acc:
        companies = [c for c in companies if str(c.get('accountant_id') or '') == acc]
    if op:
        companies = [c for c in companies if str(c.get('operator_id') or '') == op]
    if tax:
        companies = [c for c in companies if c.get('tax_system') == tax]
    if emp in ('0', '1'):
        companies = [c for c in companies if str(c.get('has_employees', 0)) == emp]
    if mil in ('0', '1'):
        companies = [c for c in companies if str(c.get('has_military', 0)) == mil]

    accountants = [dict(a) for a in accountants_raw]

    # Группировка всех компаний по сотруднику (все роли)
    all_companies_full = [dict(c) for c in companies_raw]  # без фильтра для сводки
    staff_map: dict = {}
    for c in all_companies_full:
        def _add(aid, aname, role):
            if not aid: return
            key = aid
            if key not in staff_map:
                staff_map[key] = {'id': aid, 'name': aname, 'roles': {}}
            if role not in staff_map[key]['roles']:
                staff_map[key]['roles'][role] = []
            staff_map[key]['roles'][role].append({'id': c['id'], 'name': c['name'], 'tax': c.get('tax_system','')})
        _add(c.get('accountant_id'),         c.get('accountant_name'),         'Бухгалтер')
        _add(c.get('payroll_accountant_id'),  c.get('payroll_accountant_name'), 'Зарплатник')
        _add(c.get('operator_id'),            c.get('operator_name'),           'Операционист')
        _add(c.get('hr_accountant_id'),       c.get('hr_accountant_name'),      'Кадры')

    staff_summary = sorted(staff_map.values(), key=lambda x: x['name'] or '')
    for s in staff_summary:
        s['total'] = sum(len(v) for v in s['roles'].values())

    return templates.TemplateResponse(
        request=request,
        name='dashboard.html',
        context={
            'companies':      companies,
            'accountants':    accountants,
            'staff_summary':  staff_summary,
            'today':       today.strftime('%d.%m.%Y'),
            'co_filters':  {'q': q or '', 'acc': acc or '', 'op': op or '', 'tax': tax or '', 'emp': emp or '', 'mil': mil or ''},
            'stats': {
                'companies':   len(companies),
                'accountants': len(accountants),
            },
        }
    )


MONTHS_RU = ['', 'январь', 'февраль', 'март', 'апрель', 'май', 'июнь',
             'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь']


@app.get('/company/{company_id}', response_class=HTMLResponse)
async def company_page(company_id: int, request: Request, _=Depends(require_auth)):
    company_raw = await asyncio.to_thread(db.get_company, company_id)
    if not company_raw:
        raise HTTPException(status_code=404, detail='Компания не найдена')

    company = dict(company_raw)
    today = date.today()
    accountants_raw = await asyncio.to_thread(db.get_all_accountants)

    deadlines_raw = await asyncio.to_thread(db.get_deadlines_for_company, company_id)
    works_raw     = await asyncio.to_thread(db.get_additional_works_for_company_month,
                                            company_id, today.year, today.month)

    deadlines = []
    for dl in deadlines_raw:
        d = dict(dl)
        d['due_fmt']   = fmt_date(d['due_date'])
        d['days_left'] = days_left(d['due_date'])
        deadlines.append(d)

    works = [dict(w) for w in works_raw]
    for w in works:
        w['work_date_fmt'] = fmt_date(w.get('work_date', ''))

    total_hours  = sum(w.get('hours', 0) or 0 for w in works)
    total_amount = sum(w.get('amount', 0) or 0 for w in works)

    pending  = [d for d in deadlines if d['status'] == 'pending']
    done     = [d for d in deadlines if d['status'] == 'done']
    overdue  = [d for d in deadlines if d['status'] == 'overdue']

    return templates.TemplateResponse(
        request=request,
        name='company.html',
        context={
            'company':      company,
            'pending':      pending,
            'done':         done,
            'overdue':      overdue,
            'works':        works,
            'total_hours':  total_hours,
            'total_amount': total_amount,
            'accountants':  [dict(a) for a in accountants_raw],
            'today':        today.strftime('%d.%m.%Y'),
            'today_iso':    today.isoformat(),
            'month_label':  f"{MONTHS_RU[today.month]} {today.year}",
        }
    )


# ─── Редактирование компании ──────────────────────────────────────────────────

@app.post('/companies/add')
async def company_add(
    _=Depends(require_auth),
    name: str = Form(...),
    org_type: str = Form('ООО'),
    tax_system: str = Form('УСН'),
    has_employees: str = Form('0'),
    has_military: str = Form('0'),
    accountant_id: Optional[str] = Form(None),
):
    await asyncio.to_thread(
        db.add_company,
        name=name, inn=None, tax_system=tax_system, org_type=org_type,
        has_employees=int(has_employees == '1'),
        has_military=int(has_military == '1'),
        accountant_id=int(accountant_id) if accountant_id else None,
    )
    await tg(f'🏢 <b>Добавлена новая компания</b>\n{org_type} «{name}» · {tax_system}')
    return RedirectResponse('/', status_code=303)


@app.post('/company/{company_id}/description')
async def company_description(
    company_id: int,
    _=Depends(require_auth),
    description: str = Form(''),
):
    await asyncio.to_thread(db.update_company, company_id, description=description or None)
    return RedirectResponse(f'/company/{company_id}', status_code=303)


@app.post('/company/{company_id}/delete')
async def company_delete(company_id: int, _=Depends(require_auth)):
    await asyncio.to_thread(db.deactivate_company, company_id)
    return RedirectResponse('/', status_code=303)


@app.post('/company/{company_id}/edit')
async def company_edit(
    company_id: int,
    request: Request,
    _=Depends(require_auth),
    name: str = Form(...),
    tax_system: str = Form(...),
    org_type: str = Form(...),
    has_employees: str = Form('0'),
    has_military: str = Form('0'),
    accountant_id: Optional[str] = Form(None),
    payroll_accountant_id: Optional[str] = Form(None),
    operator_id: Optional[str] = Form(None),
    hr_accountant_id: Optional[str] = Form(None),
):
    await asyncio.to_thread(
        db.update_company, company_id,
        name=name,
        tax_system=tax_system,
        org_type=org_type,
        has_employees=int(has_employees == '1'),
        has_military=int(has_military == '1'),
        accountant_id=int(accountant_id) if accountant_id else None,
        payroll_accountant_id=int(payroll_accountant_id) if payroll_accountant_id else None,
        operator_id=int(operator_id) if operator_id else None,
        hr_accountant_id=int(hr_accountant_id) if hr_accountant_id else None,
    )
    return RedirectResponse(f'/company/{company_id}', status_code=303)


# ─── Отчётность ───────────────────────────────────────────────────────────────

# ── Классификация типов отчётов ──────────────────────────────────────────────

# Квартальные (один дедлайн на квартал)
_Q_TYPES = {
    'НДС', 'Прибыль', 'Платёж Прибыль',
    'УСН', 'Платёж УСН', 'ЕСХН', 'Платёж ЕСХН',
    '6-НДФЛ', 'РСВ', 'ЕФС-1', 'БО', 'Воинский учёт', 'Статотчётность',
}
_Q_TYPE_ORDER = ['НДС', 'Прибыль', 'Платёж Прибыль', 'УСН', 'Платёж УСН',
                 'ЕСХН', 'Платёж ЕСХН', '6-НДФЛ', 'РСВ', 'ЕФС-1',
                 'Воинский учёт', 'Статотчётность', 'БО']

# Ежемесячные (отдельный столбец на каждый месяц)
_M_TYPES = {'Платёж НДС', 'Перс.сведения', 'Платёж СВ'}
_M_TYPE_ORDER = ['Перс.сведения', 'Платёж НДС', 'Платёж СВ']

_TYPE_SHORT = {
    'НДС': 'НДС', 'Прибыль': 'Прибыль', 'Платёж Прибыль': 'П.Прибыль',
    'УСН': 'УСН', 'Платёж УСН': 'П.УСН', 'ЕСХН': 'ЕСХН', 'Платёж ЕСХН': 'П.ЕСХН',
    '6-НДФЛ': '6-НДФЛ', 'РСВ': 'РСВ', 'ЕФС-1': 'ЕФС-1',
    'БО': 'БО', 'Воинский учёт': 'Воинск.', 'Статотчётность': 'Статотч.',
    'Платёж НДС': 'П.НДС', 'Перс.сведения': 'Перс.св.', 'Платёж СВ': 'П.СВ',
}

_QUARTERS = [
    (2, 'II кв. (апр–июн)',  4,  6),
    (3, 'III кв. (июл–сен)', 7,  9),
    (4, 'IV кв. (окт–дек)', 10, 12),
]

_MONTHS_RU = {
    4:'Апр', 5:'Май', 6:'Июн', 7:'Июл', 8:'Авг',
    9:'Сен', 10:'Окт', 11:'Ноя', 12:'Дек',
}


def _urgency(due: str, today_iso: str, in2: str) -> str:
    if due < today_iso: return 'overdue'
    if due <= in2:      return '2days'
    return 'ok'


def _normalize_rtype(rtype: str, rname: str) -> str:
    """Переопределяет тип для ежемесячных 'ЕФС-1' — это Перс.сведения."""
    if rtype == 'ЕФС-1' and 'Перс. сведения' in rname:
        return 'Перс.сведения'
    if rtype == 'ЕФС-1' and 'Страховые' in rname:
        return 'Платёж СВ'
    return rtype


@app.get('/reports', response_class=HTMLResponse)
async def reports_page(
    request: Request, _=Depends(require_auth),
    year: Optional[str] = None,
    acc: Optional[str] = None,
    org_type: Optional[str] = None,
):
    from datetime import timedelta
    from collections import Counter
    today     = date.today()
    yr        = int(year) if year else today.year
    today_iso = today.isoformat()
    in2       = (today + timedelta(days=2)).isoformat()

    accountants_raw = await asyncio.to_thread(db.get_all_accountants)
    companies_raw   = await asyncio.to_thread(db.get_all_companies)
    deadlines_raw   = await asyncio.to_thread(
        db.get_all_deadlines_full, yr, None, None, None, None
    )

    q2_start = f'{yr}-04-01'
    q4_end   = f'{yr}-12-31'

    def _prio(x):
        if x['status'] == 'done': return 0
        return {'overdue': 3, '2days': 2, 'ok': 1}.get(_urgency(x['due_date'], today_iso, in2), 1)

    def _build_cell(d):
        urg = 'done' if d['status'] == 'done' else _urgency(d['due_date'], today_iso, in2)
        return {'status': d['status'], 'due': d['due_date'][:10],
                'id': d['id'], 'urgency': urg, 'name': d['report_name']}

    # ── Квартальный индекс ────────────────────────────────────────────────────
    qi: dict = {}   # (cid, rtype_norm, qnum) → deadline
    # ── Ежемесячный индекс ────────────────────────────────────────────────────
    mi: dict = {}   # (cid, rtype_norm, month_num) → deadline

    for dl in deadlines_raw:
        d = dict(dl)
        due = d['due_date']
        if not (q2_start <= due <= q4_end):
            continue
        rtype = _normalize_rtype(d['report_type'], d['report_name'])
        mo = int(due[5:7])

        if rtype in _M_TYPES:
            key = (d['company_id'], rtype, mo)
            if key not in mi or _prio(d) > _prio(mi[key]):
                mi[key] = d
        elif rtype in _Q_TYPES:
            for qnum, _, qm_s, qm_e in _QUARTERS:
                if qm_s <= mo <= qm_e:
                    key = (d['company_id'], rtype, qnum)
                    if key not in qi or _prio(d) > _prio(qi[key]):
                        qi[key] = d
                    break

    # ── Столбцы квартальной матрицы ───────────────────────────────────────────
    qcol_set: set = set()
    qdue_by: dict = {}
    for (cid, rtype, qnum), d in qi.items():
        qcol_set.add((qnum, rtype))
        ck = (qnum, rtype)
        if ck not in qdue_by or d['due_date'] < qdue_by[ck]:
            qdue_by[ck] = d['due_date']

    def _qsort(c):
        q, t = c
        try: return (q, _Q_TYPE_ORDER.index(t))
        except: return (q, 99)

    q_col_defs = []
    for qnum, rtype in sorted(qcol_set, key=_qsort):
        due_lbl = ''
        raw = qdue_by.get((qnum, rtype), '')
        if raw:
            try: due_lbl = date.fromisoformat(raw).strftime('%d.%m')
            except: pass
        q_col_defs.append({
            'col_key': f'q{qnum}_{rtype}',
            'quarter_num': qnum,
            'type': rtype,
            'short': _TYPE_SHORT.get(rtype, rtype[:7]),
            'due_label': due_lbl,
        })

    q_cnt = Counter(c['quarter_num'] for c in q_col_defs)
    quarter_groups = [(qnum, ql, q_cnt[qnum]) for qnum, ql, _, _ in _QUARTERS if q_cnt[qnum]]

    # ── Столбцы ежемесячной матрицы ───────────────────────────────────────────
    # Определяем активные месяцы и типы
    m_types_present: set = set()
    m_months_present: set = set()
    for (cid, rtype, mo) in mi:
        m_types_present.add(rtype)
        m_months_present.add(mo)

    m_months = sorted(m_months_present)
    m_types  = [t for t in _M_TYPE_ORDER if t in m_types_present]

    m_col_defs = []
    for mo in m_months:
        for rtype in m_types:
            # Ищем типичную дату для заголовка
            sample = next((mi[k] for k in mi if k[1] == rtype and k[2] == mo), None)
            due_lbl = ''
            if sample:
                try: due_lbl = date.fromisoformat(sample['due_date']).strftime('%d.%m')
                except: pass
            m_col_defs.append({
                'col_key': f'm{mo}_{rtype}',
                'month_num': mo,
                'month_label': _MONTHS_RU.get(mo, str(mo)),
                'type': rtype,
                'short': _TYPE_SHORT.get(rtype, rtype[:7]),
                'due_label': due_lbl,
            })

    # Группировка по месяцам для заголовка
    m_month_groups = [(mo, _MONTHS_RU.get(mo, str(mo)), len(m_types)) for mo in m_months]

    # ── Фильтрация компаний ───────────────────────────────────────────────────
    companies = [dict(c) for c in companies_raw]
    if acc:
        companies = [c for c in companies if str(c.get('accountant_id') or '') == acc]
    if org_type:
        companies = [c for c in companies if c.get('org_type') == org_type]
    companies.sort(key=lambda c: (c.get('accountant_name') or '', c['name']))

    # ── Квартальная матрица ───────────────────────────────────────────────────
    q_matrix: dict = {}
    for c in companies:
        cid = c['id']
        row: dict = {}
        for col in q_col_defs:
            key = (cid, col['type'], col['quarter_num'])
            row[col['col_key']] = _build_cell(qi[key]) if key in qi else None
        q_matrix[str(cid)] = row

    # ── Ежемесячная матрица ───────────────────────────────────────────────────
    m_matrix: dict = {}
    for c in companies:
        cid = c['id']
        row: dict = {}
        for col in m_col_defs:
            key = (cid, col['type'], col['month_num'])
            row[col['col_key']] = _build_cell(mi[key]) if key in mi else None
        m_matrix[str(cid)] = row

    # Только компании с хотя бы одним ежемесячным дедлайном
    m_companies = [c for c in companies
                   if any(m_matrix[str(c['id'])].values())]

    # ── Счётчики ─────────────────────────────────────────────────────────────
    all_q = [v for row in q_matrix.values() for v in row.values() if v]
    all_m = [v for row in m_matrix.values() for v in row.values() if v]
    all_cells = all_q + all_m
    done_cnt    = sum(1 for v in all_cells if v['status'] == 'done')
    overdue_cnt = sum(1 for v in all_cells if v['urgency'] == 'overdue')
    critical_cnt= sum(1 for v in all_cells if v['urgency'] == '2days')
    pending_cnt = len(all_cells) - done_cnt
    pct_done    = round(done_cnt / len(all_cells) * 100) if all_cells else 0

    return templates.TemplateResponse(request=request, name='reports.html', context={
        # Квартальный блок
        'q_col_defs':      q_col_defs,
        'quarter_groups':  quarter_groups,
        'q_matrix':        q_matrix,
        # Ежемесячный блок
        'm_col_defs':      m_col_defs,
        'm_month_groups':  m_month_groups,
        'm_matrix':        m_matrix,
        'm_companies':     m_companies,
        # Общее
        'company_rows':    companies,
        'done_cnt':        done_cnt,
        'pending_cnt':     pending_cnt,
        'overdue_cnt':     overdue_cnt,
        'critical_cnt':    critical_cnt,
        'pct_done':        pct_done,
        'has_data':        bool(qi) or bool(mi),
        'accountants':     [dict(a) for a in accountants_raw],
        'year':            yr,
        'years':           list(range(today.year - 1, today.year + 2)),
        'filters':         {'acc': acc or '', 'org_type': org_type or ''},
    })


@app.get('/reports/list', response_class=HTMLResponse)
async def reports_list_page(
    request: Request, _=Depends(require_auth),
    year: Optional[str] = None,
    acc: Optional[str] = None,
    tab: Optional[str] = None,
):
    from datetime import timedelta
    today    = date.today()
    yr       = int(year) if year else today.year
    today_iso = today.isoformat()
    in2      = (today + timedelta(days=2)).isoformat()

    accountants_raw = await asyncio.to_thread(db.get_all_accountants)
    deadlines_raw   = await asyncio.to_thread(
        db.get_all_deadlines_full, yr, None,
        int(acc) if acc else None, None, None
    )

    deadlines = []
    for dl in deadlines_raw:
        d = dict(dl)
        d['due_fmt'] = fmt_date(d['due_date'])
        due = d['due_date']
        if d['status'] == 'done':
            d['urgency'] = 'done'
        elif due < today_iso:
            d['urgency'] = 'overdue'
        elif due <= in2:
            d['urgency'] = '2days'
        else:
            d['urgency'] = 'ok'
        deadlines.append(d)

    show = deadlines
    if tab == 'unsent':
        show = [d for d in deadlines if d['status'] != 'done']
    elif tab == 'critical':
        show = [d for d in deadlines if d['urgency'] in ('overdue', '2days')]

    total       = len(deadlines)
    done_cnt    = sum(1 for d in deadlines if d['status'] == 'done')
    overdue_cnt = sum(1 for d in deadlines if d['urgency'] == 'overdue')
    critical_cnt= sum(1 for d in deadlines if d['urgency'] == '2days')

    return templates.TemplateResponse(request=request, name='reports_list.html', context={
        'deadlines':    show,
        'all_count':    total,
        'done_cnt':     done_cnt,
        'pending_cnt':  total - done_cnt,
        'overdue_cnt':  overdue_cnt,
        'critical_cnt': critical_cnt,
        'pct_done':     round(done_cnt / total * 100) if total else 0,
        'accountants':  [dict(a) for a in accountants_raw],
        'today':        today.strftime('%d.%m.%Y'),
        'year':         yr,
        'years':        list(range(today.year - 1, today.year + 2)),
        'filters':      {'acc': acc or '', 'tab': tab or ''},
    })


@app.post('/reports/generate')
async def reports_generate(_=Depends(require_auth), year: int = Form(...)):
    await asyncio.to_thread(db.generate_deadlines_for_all, year)
    return RedirectResponse(f'/reports?year={year}', status_code=303)


@app.post('/reports/{deadline_id}/done')
async def report_done(deadline_id: int, _=Depends(require_auth)):
    dl = await asyncio.to_thread(db.get_deadline, deadline_id)
    await asyncio.to_thread(db.mark_deadline_done, deadline_id)
    if dl:
        d = dict(dl)
        await tg(f'📄 <b>Отчёт сдан</b>\n{d.get("company_name","?")} — {d.get("report_name","?")}')
    return RedirectResponse('/reports', status_code=303)


# ─── Дедлайны ─────────────────────────────────────────────────────────────────

@app.post('/company/{company_id}/deadline/{deadline_id}/done')
async def deadline_done(company_id: int, deadline_id: int, _=Depends(require_auth)):
    await asyncio.to_thread(db.mark_deadline_done, deadline_id)
    return RedirectResponse(f'/company/{company_id}', status_code=303)


@app.post('/company/{company_id}/deadline/add')
async def deadline_add(
    company_id: int,
    _=Depends(require_auth),
    report_name: str = Form(...),
    due_date: str = Form(...),
    period: Optional[str] = Form(None),
):
    await asyncio.to_thread(
        db.add_deadline, company_id, report_name, 'custom', due_date, period or None
    )
    return RedirectResponse(f'/company/{company_id}', status_code=303)


# ─── Доп. работы ──────────────────────────────────────────────────────────────

@app.post('/company/{company_id}/work/add')
async def work_add(
    company_id: int,
    _=Depends(require_auth),
    description: str = Form(...),
    work_type: str = Form('Прочее'),
    work_date: str = Form(...),
    hours: float = Form(0),
    amount: float = Form(0),
    accountant_id: Optional[str] = Form(None),
):
    await asyncio.to_thread(
        db.add_additional_work,
        company_id, description, work_type, work_date,
        int(accountant_id) if accountant_id else None,
        hours, amount,
    )
    return RedirectResponse(f'/company/{company_id}', status_code=303)


# ─── Бухгалтеры ───────────────────────────────────────────────────────────────

@app.get('/accountants', response_class=HTMLResponse)
async def accountants_page(request: Request, _=Depends(require_auth)):
    accs = await asyncio.to_thread(db.get_all_accountants)
    return templates.TemplateResponse(
        request=request, name='accountants.html',
        context={'accountants': [dict(a) for a in accs], 'today': date.today().strftime('%d.%m.%Y')}
    )


@app.post('/accountants/add')
async def accountant_add(
    _=Depends(require_auth),
    name: str = Form(...),
    position: Optional[str] = Form(None),
    phone: Optional[str] = Form(None),
    tg: Optional[str] = Form(None),
    is_remote: Optional[str] = Form(None),
):
    def _add():
        with db.get_db() as conn:
            conn.execute(
                'INSERT INTO accountants (name, position, phone, tg, is_remote) VALUES (?,?,?,?,?)',
                (name, position or None, phone or None, tg or None, 1 if is_remote else 0)
            )
    await asyncio.to_thread(_add)
    return RedirectResponse('/accountants', status_code=303)


@app.post('/accountants/{accountant_id}/edit')
async def accountant_edit(
    accountant_id: int,
    _=Depends(require_auth),
    name: str = Form(...),
    position: Optional[str] = Form(None),
    phone: Optional[str] = Form(None),
    email: Optional[str] = Form(None),
    tg: Optional[str] = Form(None),
    max_user_id: Optional[str] = Form(None),
    is_remote: Optional[str] = Form(None),
):
    def _update():
        with db.get_db() as conn:
            conn.execute(
                'UPDATE accountants SET name=?, position=?, phone=?, email=?, tg=?, max_user_id=?, is_remote=? WHERE id=?',
                (name, position or None, phone or None, email or None, tg or None,
                 max_user_id or None, 1 if is_remote else 0, accountant_id)
            )
    await asyncio.to_thread(_update)
    return RedirectResponse('/accountants', status_code=303)


@app.post('/accountants/{accountant_id}/delete')
async def accountant_delete(accountant_id: int, _=Depends(require_auth)):
    def _delete():
        with db.get_db() as conn:
            # Снимаем ссылки на бухгалтера в компаниях (FK ON — нельзя удалить с ссылками)
            conn.execute('UPDATE companies SET accountant_id=NULL WHERE accountant_id=?', (accountant_id,))
            conn.execute('UPDATE companies SET payroll_accountant_id=NULL WHERE payroll_accountant_id=?', (accountant_id,))
            conn.execute('UPDATE companies SET operator_id=NULL WHERE operator_id=?', (accountant_id,))
            conn.execute('DELETE FROM accountants WHERE id=?', (accountant_id,))
    await asyncio.to_thread(_delete)
    return RedirectResponse('/accountants', status_code=303)


# ─── KPI бухгалтеров ──────────────────────────────────────────────────────────

@app.get('/kpi', response_class=HTMLResponse)
async def kpi_page(
    request: Request, _=Depends(require_auth),
    year: Optional[str] = None,
    month: Optional[str] = None,
):
    today = date.today()
    sel_year  = int(year)  if year  else today.year
    sel_month = int(month) if month else today.month

    kpi_raw = await asyncio.to_thread(
        db.get_accountant_stats_full, sel_year, sel_month
    )
    kpi = []
    for s in kpi_raw:
        d = dict(s)
        total = d.get('total_deadlines') or 0
        done  = d.get('done_deadlines') or 0
        d['pct_done'] = round(done / total * 100) if total else 0
        kpi.append(d)

    MONTHS = [
        (1,'Январь'),(2,'Февраль'),(3,'Март'),(4,'Апрель'),
        (5,'Май'),(6,'Июнь'),(7,'Июль'),(8,'Август'),
        (9,'Сентябрь'),(10,'Октябрь'),(11,'Ноябрь'),(12,'Декабрь'),
    ]

    return templates.TemplateResponse(request=request, name='kpi.html', context={
        'kpi':        kpi,
        'sel_year':   sel_year,
        'sel_month':  sel_month,
        'month_label': f"{MONTHS_RU[sel_month]} {sel_year}",
        'today':      today.strftime('%d.%m.%Y'),
        'years':      list(range(today.year - 2, today.year + 2)),
        'months':     MONTHS,
    })


# ─── Дедлайны ─────────────────────────────────────────────────────────────────

@app.get('/deadlines', response_class=HTMLResponse)
async def deadlines_page(
    request: Request, _=Depends(require_auth),
    acc: Optional[str] = None, status_f: Optional[str] = None,
    period: Optional[str] = None, q: Optional[str] = None,
):
    today = date.today()
    accountants_raw = await asyncio.to_thread(db.get_all_accountants)

    # Получаем все дедлайны
    all_dl = await asyncio.to_thread(db.get_upcoming_deadlines, 365)
    overdue_dl = await asyncio.to_thread(db.get_overdue_deadlines)

    deadlines = []
    for dl in all_dl:
        d = dict(dl)
        d['due_fmt']   = fmt_date(d['due_date'])
        d['days_left'] = days_left(d['due_date'])
        deadlines.append(d)
    for dl in overdue_dl:
        d = dict(dl)
        d['due_fmt']   = fmt_date(d['due_date'])
        d['days_left'] = days_left(d['due_date'])
        if not any(x['id'] == d['id'] for x in deadlines):
            deadlines.append(d)

    # Применяем фильтры
    if acc:
        deadlines = [d for d in deadlines if str(d.get('accountant_id') or '') == acc]
    if status_f:
        if status_f == 'overdue':
            deadlines = [d for d in deadlines if d.get('status') == 'pending' and d['days_left'] < 0]
        else:
            deadlines = [d for d in deadlines if d.get('status') == status_f]
    if period:
        deadlines = [d for d in deadlines if d['due_date'][:7] == period]
    if q:
        ql = q.lower()
        deadlines = [d for d in deadlines if ql in (d.get('company_name') or '').lower()]

    deadlines.sort(key=lambda d: d['due_date'])

    return templates.TemplateResponse(request=request, name='deadlines.html', context={
        'deadlines':   deadlines,
        'accountants': [dict(a) for a in accountants_raw],
        'today':       today.strftime('%d.%m.%Y'),
        'filters':     {'acc': acc or '', 'status': status_f or '', 'period': period or today.strftime('%Y-%m'), 'q': q or ''},
    })


@app.post('/deadlines/{deadline_id}/done')
async def deadline_done_global(deadline_id: int, _=Depends(require_auth)):
    dl = await asyncio.to_thread(db.get_deadline, deadline_id)
    await asyncio.to_thread(db.mark_deadline_done, deadline_id)
    if dl:
        await tg(f'✅ <b>Дедлайн выполнен</b>\n{dict(dl).get("company_name","?")} — {dict(dl).get("report_name","?")}')
    return RedirectResponse('/deadlines', status_code=303)


# ─── Задачи ───────────────────────────────────────────────────────────────────

@app.get('/tasks', response_class=HTMLResponse)
async def tasks_page(
    request: Request, _=Depends(require_auth),
    acc: Optional[str] = None,
    status_f: Optional[str] = None,
    company: Optional[str] = None,
    priority: Optional[str] = None,
):
    today = date.today()
    accountants_raw = await asyncio.to_thread(db.get_all_accountants)
    companies_raw   = await asyncio.to_thread(db.get_all_companies)

    all_tasks_raw = await asyncio.to_thread(db.get_all_tasks)
    tasks = [dict(t) for t in all_tasks_raw]

    if acc:
        tasks = [t for t in tasks if str(t.get('accountant_id') or '') == acc]
    if company:
        tasks = [t for t in tasks if str(t.get('company_id') or '') == company]
    if priority:
        tasks = [t for t in tasks if t.get('priority') == priority]

    pending = [t for t in tasks if t.get('status') == 'pending']
    done    = [t for t in tasks if t.get('status') == 'done']
    overdue_count = sum(
        1 for t in pending
        if t.get('due_date') and t['due_date'] < today.isoformat()
    )

    return templates.TemplateResponse(request=request, name='tasks.html', context={
        'tasks':         tasks,
        'pending':       pending,
        'done':          done,
        'overdue_count': overdue_count,
        'accountants':   [dict(a) for a in accountants_raw],
        'companies':     [dict(c) for c in companies_raw],
        'today':         today.strftime('%d.%m.%Y'),
        'today_iso':     today.isoformat(),
        'filters':       {'acc': acc or '', 'status': status_f or '', 'company': company or '', 'priority': priority or ''},
    })


@app.post('/tasks/add')
async def task_add(
    _=Depends(require_auth),
    title: str = Form(...),
    accountant_id: Optional[str] = Form(None),
    company_id: Optional[str] = Form(None),
    due_date: Optional[str] = Form(None),
    priority: str = Form('normal'),
    description: Optional[str] = Form(None),
):
    await asyncio.to_thread(
        db.add_task, title,
        int(company_id) if company_id else None,
        int(accountant_id) if accountant_id else None,
        description or None, due_date or None, priority,
    )
    prio_label = {'high': '🔴', 'low': '🟢'}.get(priority, '🟡')
    await tg(f'📋 <b>Новая задача</b>\n{prio_label} {title}' + (f'\nСрок: {due_date}' if due_date else ''))
    return RedirectResponse('/tasks', status_code=303)


@app.post('/tasks/{task_id}/done')
async def task_done(task_id: int, _=Depends(require_auth)):
    t = await asyncio.to_thread(db.get_task, task_id)
    await asyncio.to_thread(db.mark_task_done, task_id)
    if t:
        await tg(f'✅ <b>Задача выполнена</b>\n{dict(t).get("title","?")}')
    return RedirectResponse('/tasks', status_code=303)


# ─── Доп. работы (сводная) ───────────────────────────────────────────────────

@app.get('/works', response_class=HTMLResponse)
async def works_page(
    request: Request, _=Depends(require_auth),
    period: Optional[str] = None,
    acc: Optional[str] = None,
    wtype: Optional[str] = None,
    q: Optional[str] = None,
    company: Optional[str] = None,
):
    from calendar_data import WORK_TYPES
    today = date.today()
    if not period:
        period = today.strftime('%Y-%m')
    year, month = int(period[:4]), int(period[5:7])

    accountants_raw = await asyncio.to_thread(db.get_all_accountants)
    companies_raw   = await asyncio.to_thread(db.get_all_companies)
    works_raw       = await asyncio.to_thread(db.get_additional_works_for_month, year, month)

    works = [dict(w) for w in works_raw]

    if acc:
        works = [w for w in works if str(w.get('accountant_id') or '') == acc]
    if company:
        works = [w for w in works if str(w.get('company_id') or '') == company]
    if wtype:
        works = [w for w in works if w.get('work_type') == wtype]
    if q:
        ql = q.lower()
        works = [w for w in works if ql in (w.get('company_name') or '').lower()]

    total_hours  = sum(w.get('hours') or 0 for w in works)
    total_amount = sum(w.get('amount') or 0 for w in works)
    companies_count = len({w['company_id'] for w in works})

    # Сводка по бухгалтерам
    by_acc: dict = {}
    for w in works:
        name = w.get('accountant_name') or '— не указан —'
        if name not in by_acc:
            by_acc[name] = {'name': name, 'cnt': 0, 'hours': 0, 'amount': 0}
        by_acc[name]['cnt']    += 1
        by_acc[name]['hours']  += w.get('hours') or 0
        by_acc[name]['amount'] += w.get('amount') or 0
    by_accountant = sorted(by_acc.values(), key=lambda x: -x['amount'])

    MONTHS_RU_GEN = ['','января','февраля','марта','апреля','мая','июня',
                     'июля','августа','сентября','октября','ноября','декабря']

    return templates.TemplateResponse(request=request, name='works.html', context={
        'works':           works,
        'total_hours':     round(total_hours, 1),
        'total_amount':    total_amount,
        'companies_count': companies_count,
        'by_accountant':   by_accountant,
        'accountants':     [dict(a) for a in accountants_raw],
        'companies':       [dict(c) for c in companies_raw],
        'work_types':      WORK_TYPES,
        'today':           today.strftime('%d.%m.%Y'),
        'today_iso':       today.isoformat(),
        'month_label':     f"{MONTHS_RU[month]} {year}",
        'filters':         {'period': period, 'acc': acc or '', 'wtype': wtype or '', 'q': q or '', 'company': company or ''},
    })


@app.post('/works/add')
async def work_add_global(
    _=Depends(require_auth),
    company_id: int = Form(...),
    accountant_id: Optional[str] = Form(None),
    work_type: str = Form(...),
    work_date: str = Form(...),
    description: str = Form(...),
    hours: float = Form(0),
    amount: float = Form(0),
):
    await asyncio.to_thread(
        db.add_additional_work,
        company_id, description, work_type, work_date,
        int(accountant_id) if accountant_id else None,
        hours, amount,
    )
    return RedirectResponse('/works', status_code=303)


# ─── Журнал ошибок ────────────────────────────────────────────────────────────

@app.get('/errors', response_class=HTMLResponse)
async def errors_page(
    request: Request, _=Depends(require_auth),
    period: Optional[str] = None, acc: Optional[str] = None,
):
    today = date.today()
    if not period:
        period = today.strftime('%Y-%m')
    year, month = int(period[:4]), int(period[5:7])

    accountants_raw = await asyncio.to_thread(db.get_all_accountants)
    companies_raw   = await asyncio.to_thread(db.get_all_companies)
    errors_raw      = await asyncio.to_thread(db.get_errors_for_month, year, month)

    errors = [dict(e) for e in errors_raw]
    if acc:
        errors = [e for e in errors if str(e.get('accountant_id') or '') == acc]

    by_accountant = {}
    for e in errors:
        n = e['accountant_name']
        by_accountant[n] = by_accountant.get(n, 0) + 1

    return templates.TemplateResponse(request=request, name='errors.html', context={
        'errors':        errors,
        'by_accountant': by_accountant,
        'accountants':   [dict(a) for a in accountants_raw],
        'companies':     [dict(c) for c in companies_raw],
        'today':         today.strftime('%d.%m.%Y'),
        'today_iso':     today.isoformat(),
        'filters':       {'period': period, 'acc': acc or ''},
    })


@app.post('/errors/add')
async def error_add(
    _=Depends(require_auth),
    accountant_id: int = Form(...),
    company_id: Optional[str] = Form(None),
    error_date: str = Form(...),
    description: str = Form(...),
):
    await asyncio.to_thread(
        db.add_error, accountant_id, description, error_date,
        int(company_id) if company_id else None,
    )
    acc = await asyncio.to_thread(db.get_accountant, accountant_id)
    acc_name = dict(acc)['name'] if acc else '?'
    await tg(f'⚠️ <b>Зафиксирована ошибка</b>\n{acc_name}\n{description[:120]}')
    return RedirectResponse('/errors', status_code=303)


# ─── Экспорт Excel ────────────────────────────────────────────────────────────

@app.get('/export/excel')
async def export_excel(_=Depends(require_auth)):
    import traceback
    try:
        return await _do_export_excel()
    except Exception as exc:
        return HTMLResponse(f'<pre>{traceback.format_exc()}</pre>', status_code=500)


async def _do_export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    today = date.today()
    companies_raw  = await asyncio.to_thread(db.get_all_companies)
    kpi_raw        = await asyncio.to_thread(db.get_accountant_stats_full, today.year, today.month)

    wb = openpyxl.Workbook()

    # ── Лист 1: Компании ──
    ws1 = wb.active
    ws1.title = 'Компании'
    hdr_font  = Font(bold=True, color='FFFFFF', size=10)
    hdr_fill  = PatternFill('solid', fgColor='1A3A5C')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = ['ID', 'Название', 'СНО', 'Тип', 'Бухгалтер', 'Сотрудники', 'Воинский учёт']
    for ci, h in enumerate(headers, 1):
        c = ws1.cell(row=1, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align
    ws1.row_dimensions[1].height = 22
    col_widths = [6, 36, 10, 8, 24, 12, 14]
    for ci, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    fill_a = PatternFill('solid', fgColor='EFF4FB')
    fill_b = PatternFill('solid', fgColor='DDEAF8')
    companies_list = list(companies_raw)
    for ri, c in enumerate(companies_list, 2):
        cd = dict(c)
        row_fill = fill_a if ri % 2 == 0 else fill_b
        vals = [cd['id'], cd['name'], cd['tax_system'], cd['org_type'],
                cd.get('accountant_name') or '',
                'Да' if cd.get('has_employees') else 'Нет',
                'Да' if cd.get('has_military') else 'Нет']
        for ci2, val in enumerate(vals, 1):
            cell = ws1.cell(row=ri, column=ci2, value=val)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical='center')
    ws1.auto_filter.ref = f'A1:G{len(companies_list)+1}'

    # ── Лист 2: KPI ──
    ws2 = wb.create_sheet('KPI бухгалтеров')
    kpi_headers = ['Бухгалтер', 'Компаний', 'Дедлайнов', 'Выполнено', 'Просрочено', '% исп.', 'Ошибок', 'Доп.часы', 'Доп.сумма']
    for ci, h in enumerate(kpi_headers, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align
    ws2.row_dimensions[1].height = 22
    kpi_widths = [26, 10, 12, 12, 12, 10, 10, 12, 14]
    for ci, w in enumerate(kpi_widths, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    for ri, s in enumerate(kpi_raw, 2):
        sd = dict(s)
        total = sd.get('total_deadlines') or 0
        done  = sd.get('done_deadlines') or 0
        pct   = round(done / total * 100) if total > 0 else 0
        row_fill = fill_a if ri % 2 == 0 else fill_b
        vals = [sd['name'], sd.get('company_count') or 0, total, done,
                sd.get('overdue_deadlines') or 0, f'{pct}%',
                sd.get('error_count') or 0,
                sd.get('extra_hours') or 0,
                sd.get('extra_amount') or 0]
        for ci2, val in enumerate(vals, 1):
            cell = ws2.cell(row=ri, column=ci2, value=val)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical='center')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename_ascii = f'IF_export_{today.strftime("%Y-%m-%d")}.xlsx'
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename_ascii}"'},
    )


@app.get('/tg/test')
async def tg_test(_=Depends(require_auth)):
    if not TG_TOKEN or not TG_CHAT:
        return {'status': 'error', 'message': 'TELEGRAM_BOT_TOKEN или TELEGRAM_CHAT_ID не заполнены в .env'}
    await tg('🔔 <b>Тест уведомлений</b>\nCRM Империя Финанс работает!\n✅ Уведомления настроены.')
    return {'status': 'ok', 'message': 'Сообщение отправлено в Telegram'}


if __name__ == '__main__':
    print(f'Дашборд: http://localhost:{WEB_PORT}  |  логин: admin / {WEB_PASSWORD}')
    uvicorn.run('web_app:app', host='0.0.0.0', port=WEB_PORT, reload=False)
